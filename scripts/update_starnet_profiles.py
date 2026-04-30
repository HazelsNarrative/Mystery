#!/usr/bin/env python3
import csv
import html
import json
import os
import re
from collections import defaultdict
from pathlib import Path

from bs4 import BeautifulSoup, NavigableString
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "starnet"
XLSX_PATH = DATA_DIR / "starnet.xlsx"


def read_csv(path):
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_sheet_from_xlsx(path, sheet_name):
    if not path.exists():
        return []
    wb = load_workbook(filename=str(path), data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    data = []
    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            value = row[idx] if idx < len(row) else ""
            item[key] = "" if value is None else str(value)
        data.append(item)
    return data


def text_to_html(text):
    safe = html.escape((text or "").strip())
    return safe.replace("\n", "<br>")


def decorate_tags(text):
    # Keep StarNet tag style while still allowing free text input.
    return re.sub(
        r"(#[^\s#]+)",
        lambda m: f'<span class="tag">{html.escape(m.group(1))}</span>',
        text_to_html(text),
    )


def stable_num(seed_text, lower, upper):
    span = upper - lower + 1
    return lower + (abs(hash(seed_text)) % span)


def resolve_asset_href(asset_path: str, file_ref: str) -> str:
    """assets/... → 相对于当前页的 URL。"""
    ap = (asset_path or "").strip()
    fr = (file_ref or "").strip()
    if not ap or re.match(r"^(https?:)?//", ap):
        return ap
    if not fr:
        return ap
    page_dir = (ROOT / fr).resolve().parent
    image_abs = (ROOT / ap).resolve()
    return os.path.relpath(str(image_abs), str(page_dir)).replace("\\", "/")


def make_comment_username(profile_id, post_order, comment_order):
    base = abs(hash(f"{profile_id}:{post_order}:{comment_order}")) % 10000
    return f"StarNet{base:04d}"


def normalize_profile_type(value):
    raw = (value or "").strip().lower()
    if raw in {"artist", "艺人"}:
        return "artist"
    if raw in {"fan", "fans", "粉丝", "素人", "civilian"}:
        return "fan"
    if raw in {"hater", "haters", "黑粉"}:
        return "hater"
    return "fan"


def profile_type_to_dir(profile_type):
    if profile_type == "artist":
        return "social/artists"
    if profile_type == "hater":
        return "social/civilians/haters"
    return "social/civilians/fans"


def slugify_ascii(value, fallback):
    s = re.sub(r"[^a-zA-Z0-9]+", "_", (value or "").strip().lower()).strip("_")
    return s or fallback


def slug_from_profile(profile_row: dict, file_path: str) -> str:
    raw = (profile_row.get("profile_slug") or "").strip()
    s = re.sub(r"[^a-zA-Z0-9_]+", "_", raw).strip("_")
    if s:
        return s
    base = Path(file_path).stem.removeprefix("starnet-social-")
    base = re.sub(r"[^a-zA-Z0-9_]+", "_", base).strip("_")
    return base or slugify_ascii(profile_row.get("profile_id") or "user", "user")


def apply_profile_identity_urls(content: str, handle_slug: str) -> str:
    url_lit = html.escape(f"https://starnet.social/{handle_slug}")
    content = replace_or_fail(
        r"<title>[^<]*StarNet[^<]*</title>",
        f"<title>@{handle_slug} | StarNet</title>",
        content,
        "未找到 title 区块",
    )
    content = replace_or_fail(
        r'(<p class="url">)(.*?)(</p>)',
        lambda m: f"{m.group(1)}{url_lit}{m.group(3)}",
        content,
        "未找到 url 行",
        count=1,
    )
    content = replace_or_fail(
        r'(<div class="profile">[\s\S]*?<p class="id">)(@)([^<]+)(</p>)',
        lambda m: f"{m.group(1)}@{html.escape(handle_slug)}{m.group(4)}",
        content,
        "未找到 profile id",
    )
    return content


def linkify_blackhole_in_comment(fragment_html: str, file_path: str) -> str:
    if "blackhole.html" not in fragment_html:
        return fragment_html
    page_dir = (ROOT / file_path).resolve().parent
    target = (ROOT / "blackhole.html").resolve()
    href = os.path.relpath(str(target), str(page_dir)).replace("\\", "/")
    anchor = (
        f'<a href="{html.escape(href)}" target="_blank" rel="noopener noreferrer">'
        f"blackhole.html</a>"
    )
    return fragment_html.replace("blackhole.html", anchor, 1)


def merge_artist_feed_panel(html_fragment: str, rebuilt_html: str) -> str:
    soup = BeautifulSoup(html_fragment, "html.parser")
    panel = soup.select_one("section.panel") or soup.select_one(".panel")
    if not panel:
        raise ValueError("未找到 section.panel")

    for art in panel.select("article.post"):
        art.decompose()

    for box in panel.select("#protectedContent"):
        box.decompose()

    holder_soup = BeautifulSoup(f"<div>{rebuilt_html}</div>", "html.parser")
    holder = holder_soup.div
    for node in list(holder.children):
        if isinstance(node, NavigableString) and not node.strip():
            continue
        panel.append(node)

    return str(soup)


def build_new_profile_html(file_path, display_name, bio):
    rel_prefix = "../" * (len(Path(file_path).parts) - 1)
    logo_path = f"{rel_prefix}assets/starnet-logo.png"
    bg_path = f"{rel_prefix}assets/starnet-home-bg.png"
    profile_name = display_name or "未命名用户"
    profile_id = "@" + slugify_ascii(display_name, "new_user")
    profile_url = f"https://starnet.social/{slugify_ascii(display_name, 'new_user')}"
    profile_bio = bio or "这个人很懒，还没有写简介。"
    return f"""<!DOCTYPE html>
<html lang="zh">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{html.escape(profile_id)} | StarNet</title>
  <style>
    :root {{ --bg:#f8f9fb; --card:#fff; --line:#e6ecf5; --text:#0f1419; --sub:#536471; --brand:#6f77ff; }}
    * {{ box-sizing:border-box; }} body {{ margin:0; font-family:"PingFang SC","Microsoft YaHei",Arial,sans-serif; background:var(--bg); color:var(--text); }}
    .app {{ width:min(900px,94vw); margin:0 auto; padding:20px 0 30px; }} .panel {{ background:var(--card); border:1px solid var(--line); border-radius:18px; overflow:hidden; box-shadow:0 10px 24px rgba(63,77,110,0.08); }}
    .site-head {{ padding:12px 16px; border-bottom:1px solid var(--line); font-size:13px; color:var(--sub); display:flex; justify-content:space-between; align-items:center; gap:10px; }}
    .brand-logo {{ width:100px; height:auto; display:block; transform:scale(1.62); transform-origin:left center; margin-left:-10px; }}
    .url {{ margin:0; font-size:12px; color:var(--sub); word-break:break-all; }}
    .cover {{ height:200px; position:relative; background-image:linear-gradient(120deg, rgba(149,172,210,0.76), rgba(112,142,189,0.75) 45%, rgba(77,103,154,0.8)), url("{bg_path}"); background-size:cover; background-position:center; }}
    .avatar {{ position:absolute; left:18px; bottom:-34px; width:84px; height:84px; border-radius:50%; border:4px solid #fff; box-shadow:0 8px 18px rgba(45,77,98,.25); display:grid; place-items:center; font-weight:800; color:#284661; background:radial-gradient(circle at 30% 30%,#fff,#cde1ec 45%,#779bb5); }}
    .profile {{ padding:46px 18px 14px; border-bottom:1px solid var(--line); }} .name {{ margin:0; font-size:22px; font-weight:800; }} .id {{ margin:4px 0 10px; color:var(--sub); font-size:13px; }}
    .bio {{ margin:0; line-height:1.7; color:#243046; font-size:14px; }} .stats {{ margin-top:10px; display:flex; gap:16px; font-size:12px; color:var(--sub); }} .stats strong {{ color:#263a61; }}
    .post {{ padding:14px 18px; border-bottom:1px solid var(--line); }} .post:last-child {{ border-bottom:none; }} .post-head {{ font-size:13px; color:var(--sub); margin-bottom:6px; }} .post p {{ margin:0; font-size:14px; line-height:1.75; color:#1d2432; }}
    .post-stats {{ margin-top:8px; font-size:12px; color:#6f7f94; display:flex; gap:14px; }}
    .post-public {{ background:#fff; }} .post-followers {{ background:linear-gradient(180deg, #fcf8ff, #f8f4ff); border-left:3px solid #9d8dff; }}
    .follow-btn {{ margin-top:10px; border:none; border-radius:999px; padding:8px 14px; font-size:12px; font-weight:700; color:#fff; background:linear-gradient(90deg,#8a6dff,#4f81df); cursor:pointer; }}
    .follow-btn.following {{ background:#d1d5db; color:#4b5563; border:1px solid #9ca3af; }}
    .locked-tip {{ margin:12px 18px 14px; padding:10px 12px; border:1px solid #e2d6ff; border-radius:12px; background:#faf7ff; color:#6a5f8a; font-size:12px; line-height:1.7; }}
    .modal {{ position:fixed; inset:0; display:none; align-items:center; justify-content:center; background:rgba(20,24,40,.45); z-index:200; }}
    .modal.show {{ display:flex; }}
    .modal-card {{ width:min(360px,92vw); background:#fff; border-radius:14px; border:1px solid #e1e8f4; padding:16px; }}
    .modal-title {{ margin:0 0 8px; font-size:15px; color:#2c3f63; font-weight:800; }}
    .modal-question {{ margin:0 0 10px; font-size:13px; color:#5f6f89; }}
    .modal-input {{ width:100%; border:1px solid #d6dfef; border-radius:10px; padding:9px 10px; font-size:13px; margin-bottom:8px; }}
    .modal-error {{ min-height:1.2em; margin:0 0 8px; font-size:12px; color:#d44c72; }}
    .modal-actions {{ display:flex; justify-content:flex-end; gap:8px; }}
    .modal-actions button {{ border:none; border-radius:999px; padding:7px 12px; font-size:12px; font-weight:700; cursor:pointer; }}
    .btn-cancel {{ background:#eef2fb; color:#5f6f89; }}
    .btn-ok {{ background:linear-gradient(90deg,#8a6dff,#4f81df); color:#fff; }}
  </style>
</head>
<body>
  <main class="app"><section class="panel">
    <div class="site-head"><img class="brand-logo" src="{logo_path}" alt="StarNet logo"><p class="url">{html.escape(profile_url)}</p></div><div class="cover"><div class="avatar">星</div></div>
    <div class="profile"><h2 class="name">{html.escape(profile_name)}</h2><p class="id">{html.escape(profile_id)}</p><p class="bio">{html.escape(profile_bio)}</p><button class="follow-btn" id="followBtn" type="button">关注</button><div class="stats"><span><strong>0</strong> 关注</span><span><strong>0</strong> 粉丝</span><span><strong>0</strong> 动态</span></div></div>
    <p class="locked-tip" id="lockedTip">该账号已开启访问权限，关注并通过验证后可查看全部动态。</p>
    <div id="protectedContent" hidden></div>
  <div class="modal" id="followModal">
    <div class="modal-card">
      <p class="modal-title">关注验证</p>
      <p class="modal-question">请输入通关口令</p>
      <input class="modal-input" id="answerInput" type="text" placeholder="请输入答案">
      <p class="modal-error" id="answerError"></p>
      <div class="modal-actions">
        <button class="btn-cancel" type="button" id="btnCancel">取消</button>
        <button class="btn-ok" type="button" id="btnSubmit">提交</button>
      </div>
    </div>
  </div>
  <script>
    (function () {{
      var followBtn = document.getElementById("followBtn");
      var followModal = document.getElementById("followModal");
      var answerInput = document.getElementById("answerInput");
      var answerError = document.getElementById("answerError");
      var btnCancel = document.getElementById("btnCancel");
      var btnSubmit = document.getElementById("btnSubmit");
      var protectedContent = document.getElementById("protectedContent");
      var lockedTip = document.getElementById("lockedTip");
      var expected = "1234";
      var isFollowing = false;
      var profileIdEl = document.querySelector(".profile .id");
      var urlEl = document.querySelector(".url");
      var profileId = (profileIdEl && String(profileIdEl.textContent || "").trim()) || (urlEl && String(urlEl.textContent || "").trim()) || String(location.pathname || location.href);
      var followStorageKey = "starnet_follow_state::" + profileId;

      function persistFollowState() {{
        try {{
          localStorage.setItem(followStorageKey, isFollowing ? "1" : "0");
        }} catch (e) {{}}
      }}

      function loadFollowState() {{
        try {{
          isFollowing = localStorage.getItem(followStorageKey) === "1";
        }} catch (e) {{
          isFollowing = false;
        }}
      }}
      function render() {{
        followBtn.textContent = isFollowing ? "已关注" : "关注";
        followBtn.classList.toggle("following", isFollowing);
        if (protectedContent && protectedContent.querySelector(".post.post-followers")) {{
          protectedContent.hidden = !isFollowing;
        }} else {{
          document.querySelectorAll(".post.post-followers").forEach(function (el) {{
            el.hidden = !isFollowing;
          }});
          if (protectedContent) {{
            protectedContent.hidden = true;
          }}
        }}
        lockedTip.hidden = isFollowing;
      }}
      function closeModal() {{
        followModal.classList.remove("show");
        answerInput.value = "";
        answerError.textContent = "";
      }}
      followBtn.addEventListener("click", function () {{
        if (isFollowing) {{
          isFollowing = false;
          persistFollowState();
          render();
          return;
        }}
        followModal.classList.add("show");
      }});
      btnCancel.addEventListener("click", closeModal);
      btnSubmit.addEventListener("click", function () {{
        if (String(answerInput.value || "").trim() !== expected) {{
          answerError.textContent = "答案错误，请重试。";
          return;
        }}
        isFollowing = true;
        persistFollowState();
        closeModal();
        render();
      }});
      followModal.addEventListener("click", function (e) {{
        if (e.target === followModal) closeModal();
      }});
      loadFollowState();
      render();
    }})();
  </script>
</body>
</html>
"""


def resolve_file_path(profile_id, profile_row):
    file_path = (profile_row.get("file") or "").strip()
    if file_path:
        return file_path
    profile_type = normalize_profile_type(profile_row.get("profile_type"))
    file_dir = profile_type_to_dir(profile_type)
    slug = (profile_row.get("profile_slug") or "").strip()
    if not slug:
        slug = slugify_ascii(profile_id, f"profile_{profile_id.lower()}")
    return f"{file_dir}/starnet-social-{slug}.html"


def needs_profile_html_seed(abs_path: Path) -> bool:
    """空文件或以残片开头的 HTML（无 DOCTYPE）无法走身份与动态注入逻辑，须写入 StarNet profile 骨架。"""
    if not abs_path.exists():
        return True
    try:
        if abs_path.stat().st_size < 200:
            return True
    except OSError:
        return True
    try:
        head = abs_path.read_text(encoding="utf-8", errors="replace").lstrip("\ufeff \t\r\n")[:24000]
    except OSError:
        return True
    return not re.search(r"<!DOCTYPE\s+html\b", head, flags=re.I)


def ensure_profile_file(profile_id, file_path, profile_row):
    abs_path = ROOT / file_path
    if not needs_profile_html_seed(abs_path):
        return
    abs_path.parent.mkdir(parents=True, exist_ok=True)
    display_name = (profile_row.get("display_name") or "").strip()
    bio = (profile_row.get("bio") or "").strip()
    abs_path.write_text(
        build_new_profile_html(file_path=file_path, display_name=display_name, bio=bio),
        encoding="utf-8",
    )


def extract_comments_from_post(post_row):
    comments = []
    for i in range(1, 9):
        nick = (post_row.get(f"comment_nickname_{i}") or "").strip()
        text = (post_row.get(f"comment_text_{i}") or "").strip()
        img = (post_row.get(f"comment_image_{i}") or "").strip()
        img_alt = (post_row.get(f"comment_image_alt_{i}") or "").strip()
        if not text:
            continue
        comments.append(
            {
                "nickname": nick,
                "text": text,
                "order": i,
                "image": img,
                "image_alt": img_alt or "表情包",
            }
        )
    return comments


def fmt_cn_wan(value: float) -> str:
    s = f"{value:.1f}".rstrip("0").rstrip(".")
    return f"{s}万"


def format_artist_engagement(raw: int) -> str:
    if raw <= 0:
        return "0"
    return fmt_cn_wan(raw / 10000.0)


def build_stats(
    profile_id,
    post_order,
    is_followers_post,
    comment_count,
    is_artist=False,
    civilian_fans_style=False,
    civilian_constell_hater_heavy=False,
):
    seed = f"{profile_id}:{post_order}:{1 if is_followers_post else 0}"
    if is_artist:
        hot_repost = stable_num(seed + ":hot_rp", 0, 99) >= 82
        if hot_repost:
            reposts_raw = stable_num(seed + ":rhi", 102_000, 186_500)
        else:
            reposts_raw = stable_num(seed + ":r", 12_600, 98_900)
        likes_raw = stable_num(seed + ":l", 29_700, 198_900)
        comments_floor = stable_num(seed + ":c0", max(13_900, comment_count + 6900), 97_900)
        comments_raw = max(comments_floor, comment_count + 5800)

        return (
            format_artist_engagement(reposts_raw),
            format_artist_engagement(likes_raw),
            format_artist_engagement(comments_raw),
        )

    # @constellhater：大号黑粉体感，四五万粉；转发偏少，赞评明显更高（仍非艺人量级）
    if civilian_fans_style and civilian_constell_hater_heavy:
        reposts = stable_num(seed + ":r", 28, 180)
        likes = stable_num(seed + ":l", 8200, 38800)
        floor = stable_num(seed + ":c0", comment_count + 180, comment_count + 980)
        comments = max(floor, comment_count + stable_num(seed + ":cx", 620, 4200))
        return str(reposts), str(likes), str(comments)

    # 素人小号（粉丝/黑粉主页）：转发几乎可忽略，点赞多位数以内，评论与真实楼层一致量级
    if civilian_fans_style:
        reposts = stable_num(seed + ":r", 0, 6)
        likes = stable_num(seed + ":l", 3, 99)
        comment_extra = stable_num(seed + ":c", 0, 9)
        comments = min(99, max(comment_count, comment_extra))
        return str(reposts), str(likes), str(comments)

    if is_followers_post:
        reposts = stable_num(seed + ":r", 20, 480)
        likes = stable_num(seed + ":l", 300, 1600)
        comments = max(comment_count, stable_num(seed + ":c", 120, 980))
    else:
        reposts = stable_num(seed + ":r", 600, 5800)
        likes = stable_num(seed + ":l", 2000, 36000)
        comments = max(comment_count, stable_num(seed + ":c", 180, 3200))
    return str(reposts), str(likes), str(comments)


def linkify_onboarding_starnet_portal(fragment_html: str, file_path: str) -> str:
    """表中「入驻…starnet.social」常为纯文本，艺人动态补回站点入口链接。"""
    norm = file_path.replace("\\", "/")
    if "/artists/" not in norm:
        return fragment_html
    if "starnet.social" not in fragment_html or "入驻" not in fragment_html:
        return fragment_html
    if re.search(r"<a\b[^>]*>\s*starnet\.social\s*</a>", fragment_html):
        return fragment_html
    page_dir = (ROOT / file_path).resolve().parent
    target = (ROOT / "social" / "starnet-home.html").resolve()
    href = os.path.relpath(str(target), str(page_dir)).replace("\\", "/")
    anchor = f'<a href="{html.escape(href)}">starnet.social</a>'
    return fragment_html.replace("starnet.social", anchor, 1)


def patch_civilian_profile_stats_bar(content: str, profile_id: str, feed_count: int, file_path: str) -> str:
    """素人主页顶栏「关注／粉丝／动态」：小号量级，与同目录粉丝页观感一致。"""
    norm = file_path.replace("\\", "/")
    if "/civilians/fans/" not in norm and "/civilians/haters/" not in norm:
        return content
    seed = f"{profile_id}::civilian_profile_banner"
    following = stable_num(seed + ":following", 28, 198)
    if "/civilians/haters/" in norm:
        if "constellhater" in norm:
            followers = stable_num(seed + ":followers", 40_500, 50_900)
        else:
            followers = stable_num(seed + ":followers", 320, 1680)
    else:
        followers = stable_num(seed + ":followers", 320, 5200)
    if "/civilians/haters/" in norm and "constellhater" in norm:
        # 与稿号【20011】等叙事一致：bot 累计动态至少已过两万
        dyn_lo = 20011 + stable_num(seed + ":d_lo", 0, 3200)
        dyn_hi = dyn_lo + stable_num(seed + ":d_hi", 600, 7800)
        dynamics = stable_num(seed + ":dyn", dyn_lo, dyn_hi)
    else:
        dyn_lo = max(12, feed_count + stable_num(seed + ":d_lo", 2, 18))
        dyn_hi = max(dyn_lo + 8, feed_count + stable_num(seed + ":d_hi", 24, 120))
        dynamics = stable_num(seed + ":dyn", dyn_lo, dyn_hi)

    stats_html = (
        f'<div class="stats"><span><strong>{following}</strong> 关注</span>'
        f'<span><strong>{followers}</strong> 粉丝</span>'
        f'<span><strong>{dynamics}</strong> 动态</span></div>'
    )
    return replace_or_fail(
        r"<div\s+class=\"stats\">[\s\S]*?</div>",
        stats_html,
        content,
        "未找到首页统计栏 .stats（素人）",
        count=1,
    )


def patch_artist_profile_stats_bar(content: str, profile_id: str, feed_count: int) -> str:
    seed = f"{profile_id}::banner"
    following = stable_num(seed + ":following", 56, 412)
    # P004 = CONSTELL-Regulus：粉丝数约四千万（其余艺人保持原区间）
    if profile_id == "P004":
        fans_raw = stable_num(seed + ":fans", 39_200_000, 40_800_000)
    else:
        fans_raw = stable_num(seed + ":fans", 22_400_000, 51_200_000)
    fans_disp = fmt_cn_wan(fans_raw / 10000.0)
    dynamics = stable_num(seed + ":dyn", max(620, feed_count + 460), feed_count + 3580)

    stats_html = (
        f'<div class="stats"><span><strong>{following}</strong> 关注</span>'
        f'<span><strong>{fans_disp}</strong> 粉丝</span>'
        f'<span><strong>{dynamics:,}</strong> 动态</span></div>'
    )

    return replace_or_fail(
        r"<div\s+class=\"stats\">[\s\S]*?</div>",
        stats_html,
        content,
        "未找到首页统计栏 .stats（艺人）",
        count=1,
    )


def build_post_html(profile_id, display_name, post_row):
    file_ref = post_row.get("file") or ""
    norm_ref = file_ref.replace("\\", "/")
    is_artist = "/social/artists/" in norm_ref or "/artists/" in norm_ref
    civilian_fans_style = "/civilians/fans/" in norm_ref or "/civilians/haters/" in norm_ref
    civilian_constell_hater_heavy = (
        civilian_fans_style and "constellhater" in norm_ref
    )
    followers_only = (post_row.get("followers_only") or "").strip()
    is_followers_post = followers_only == "1"
    classes = "post post-followers" if is_followers_post else "post post-public"
    head = f"@{html.escape(display_name)}· {(post_row.get('time') or '').strip()}"
    content = decorate_tags(post_row.get("text") or "")
    content = linkify_onboarding_starnet_portal(content, file_ref)
    image_path = (post_row.get("image") or "").strip()
    image_alt = (post_row.get("image_alt") or "动态配图").strip()
    comments = extract_comments_from_post(post_row)
    reposts, likes, comment_count_display = build_stats(
        profile_id=profile_id,
        post_order=(post_row.get("post_order") or "").strip(),
        is_followers_post=is_followers_post,
        comment_count=len(comments),
        is_artist=is_artist,
        civilian_fans_style=civilian_fans_style,
        civilian_constell_hater_heavy=civilian_constell_hater_heavy,
    )

    parts = [
        f'<article class="{classes}">',
        f'<div class="post-head">{head}</div>',
        f"<p>{content}</p>",
    ]

    if image_path:
        image_src = resolve_asset_href(image_path, post_row.get("file") or "")
        parts.append(
            '<img class="post-image zoomable" '
            f'src="{html.escape(image_src)}" '
            f'alt="{html.escape(image_alt)}">'
        )

    parts.extend(
        [
        (
            '<div class="post-stats">'
            f"<span>转发 {reposts}</span>"
            f"<span>赞 {likes}</span>"
            f"<span>评论 {comment_count_display}</span>"
            "</div>"
        ),
        ]
    )

    if comments:
        parts.append('<div class="comments">')
        for c in comments:
            idx = int(c.get("order") or 0)
            username = (c.get("nickname") or "").strip()
            if not username:
                username = make_comment_username(
                    profile_id, post_row.get("post_order", ""), idx
                )
            content_text = text_to_html(c.get("text") or "")
            content_text = linkify_blackhole_in_comment(
                content_text, post_row.get("file") or ""
            )
            parts.append(
                f'<p class="comment"><strong>{html.escape(username)}：</strong>{content_text}</p>'
            )
            cimg = (c.get("image") or "").strip()
            if cimg:
                cis = resolve_asset_href(cimg, post_row.get("file") or "")
                ialt = (c.get("image_alt") or "表情包").strip()
                parts.append(
                    '<img class="comment-media zoomable" '
                    'style="max-width:240px;display:block;margin-top:8px;border-radius:12px;'
                    'border:1px solid #d8e1ee;cursor:zoom-in" '
                    f'src="{html.escape(cis)}" '
                    f'alt="{html.escape(ialt)}">'
                )
        parts.append("</div>")

    parts.append("</article>")
    return "".join(parts), is_followers_post


def replace_or_fail(pattern, replacement, content, flag_msg, count=0):
    kwargs = dict(flags=re.S)
    if count:
        kwargs["count"] = count
    new_content, replaced = re.subn(pattern, replacement, content, **kwargs)
    if replaced == 0:
        raise ValueError(flag_msg)
    return new_content


_FAN_IIFE_NEEDLE = '(function () {\n      var followBtn'
_FAN_IIFE_NEEDLE_CRLF = '(function () {\r\n      var followBtn'


_FAN_CHROME_CSS = """
    /* starnet-fan-chrome */
    .tag { color:var(--brand); }
    .post-stats { align-items:center; flex-wrap:wrap; }
    .comments { margin-top:10px; border-top:1px dashed #d8e1ee; padding-top:8px; }
    .comments[hidden] { display:none; }
    .comment { margin:0 0 6px; font-size:12px; color:#4c5e78; line-height:1.6; }
    .comment strong { color:#2a426a; }
    .comment-toggle { margin-left:auto; border:none; background:transparent; color:#74839a; font-size:12px; line-height:1.4; padding:0; cursor:pointer; }
    .comment-toggle:hover { color:#5c6d86; }
    .comment-toggle:focus-visible { outline:1px solid #c8d4e7; outline-offset:2px; border-radius:6px; }
    .comment-media { max-width:240px; margin-top:8px; border-radius:12px; border:1px solid #d8e1ee; display:block; cursor:zoom-in; }
    .post-image { margin:10px auto 0; width:auto; height:auto; max-width:280px; max-height:min(560px,80vh); border:1px solid #d4ddee; border-radius:12px; display:block; cursor:zoom-in; }
    .image-lightbox { position:fixed; inset:0; display:none; align-items:center; justify-content:center; background:rgba(12,16,28,.82); z-index:260; padding:24px; }
    .image-lightbox.show { display:flex; }
    .image-lightbox img { max-width:min(1100px,92vw); max-height:88vh; border-radius:12px; border:1px solid rgba(255,255,255,.35); box-shadow:0 20px 45px rgba(0,0,0,.45); background:#111827; }
"""


_FAN_SETUP_HOOK = """      function setupImageZoom() {
        var lightbox = document.getElementById("imageLightbox");
        var lightboxImage = document.getElementById("lightboxImage");
        if (!lightbox || !lightboxImage) return;
        document.querySelectorAll(".post img, .comments img").forEach(function (img) {
          img.addEventListener("click", function () {
            lightboxImage.src = img.currentSrc || img.src;
            lightboxImage.alt = img.alt || "查看大图";
            lightbox.classList.add("show");
            lightbox.setAttribute("aria-hidden", "false");
          });
        });
        lightbox.addEventListener("click", function (e) {
          if (e.target !== lightbox && e.target !== lightboxImage) return;
          lightbox.classList.remove("show");
          lightboxImage.removeAttribute("src");
          lightbox.setAttribute("aria-hidden", "true");
        });
      }
      function setupCommentToggles() {
        document.querySelectorAll(".post").forEach(function (post) {
          var comments = post.querySelector(".comments");
          if (!comments || !comments.querySelector(".comment")) return;
          comments.hidden = true;
          var toggleBtn = document.createElement("button");
          toggleBtn.type = "button";
          toggleBtn.className = "comment-toggle";
          function syncToggleLabel() {
            var isExpanded = !comments.hidden;
            toggleBtn.textContent = isExpanded ? "收起评论 ▴" : "评论 ▾";
            toggleBtn.setAttribute("aria-expanded", isExpanded ? "true" : "false");
          }
          toggleBtn.addEventListener("click", function () {
            comments.hidden = !comments.hidden;
            syncToggleLabel();
          });
          syncToggleLabel();
          var stats = post.querySelector(".post-stats");
          if (stats) {
            stats.appendChild(toggleBtn);
            return;
          }
          var postBody = post.querySelector("p");
          if (postBody) {
            postBody.insertAdjacentElement("afterend", toggleBtn);
          } else {
            post.appendChild(toggleBtn);
          }
        });
      }
"""


def ensure_fan_page_chrome(html: str, file_path: str) -> str:
    """civilians 下的粉丝/黑粉主页：评论区折叠、统计栏上的「评论 ▾」、配图 lightbox（与现有粉丝页一致）。"""
    norm = file_path.replace("\\", "/")
    if "/civilians/fans/" not in norm and "/civilians/haters/" not in norm:
        return html

    def _inject_image_lightbox(s: str) -> str:
        head_before_script = (s.split("<script>", 1)[0] if "<script>" in s else s)
        if '<div class="image-lightbox"' in head_before_script:
            return s
        m = re.search(r"(?m)^(?P<ind>[ \t]*)(<div class=\"modal\" id=\"followModal\">)", s)
        if not m:
            return s
        ind = m.group("ind")
        block = (
            f'{ind}  <!-- starnet-fan-chrome -->\n'
            f'{ind}<div class="image-lightbox" id="imageLightbox" aria-hidden="true">\n'
            f'{ind}  <img id="lightboxImage" alt="查看大图">\n'
            f'{ind}</div>\n'
            f'{m.group(0)}'
        )
        return s[: m.start()] + block + s[m.end() :]

    needle_ok = _FAN_IIFE_NEEDLE in html or _FAN_IIFE_NEEDLE_CRLF in html
    if "setupCommentToggles" in html and not needle_ok:
        return _inject_image_lightbox(html)

    if not needle_ok:
        return html

    if "setupCommentToggles" in html:
        return _inject_image_lightbox(html)

    style_close = "</style>"
    ix = html.find(style_close)
    if ix != -1 and "/* starnet-fan-chrome */" not in html:
        html = html[:ix] + _FAN_CHROME_CSS + "\n  " + html[ix:]

    html = _inject_image_lightbox(html)

    if _FAN_IIFE_NEEDLE in html:
        html = html.replace(_FAN_IIFE_NEEDLE, "(function () {\n" + _FAN_SETUP_HOOK + "      var followBtn", 1)
    elif _FAN_IIFE_NEEDLE_CRLF in html:
        hk = _FAN_SETUP_HOOK.replace("\n", "\r\n")
        html = html.replace(_FAN_IIFE_NEEDLE_CRLF, "(function () {\r\n" + hk + "      var followBtn", 1)

    html, nrep = re.subn(
        r"render\(\);\s*\r?\n\s*\}\)\(\);",
        "render();\n      setupCommentToggles();\n      setupImageZoom();\n    })();",
        html,
        count=1,
    )
    if nrep != 1:
        raise ValueError(f"无法在粉丝主页注入脚本结尾（需含 render(); 后紧跟 IIFE 结束）: {file_path}")

    return html


_HATER_FOLLOW_STORAGE_PREFIX = "starnet_follow_state::hater_v2::"
_FAN_FOLLOW_STORAGE_PREFIX = "starnet_follow_state::"


def _follow_local_storage_snippet(storage_prefix_inside: str) -> str:
    return f"""
      var profileIdEl = document.querySelector(".profile .id");
      var urlEl = document.querySelector(".url");
      var profileId = (profileIdEl && String(profileIdEl.textContent || "").trim()) || (urlEl && String(urlEl.textContent || "").trim()) || String(location.pathname || location.href);
      var followStorageKey = "{storage_prefix_inside}" + profileId;

      function persistFollowState() {{
        try {{
          localStorage.setItem(followStorageKey, isFollowing ? "1" : "0");
        }} catch (e) {{}}
      }}

      function loadFollowState() {{
        try {{
          isFollowing = localStorage.getItem(followStorageKey) === "1";
        }} catch (e) {{
          isFollowing = false;
        }}
      }}
"""


_FOLLOW_LOCAL_STORAGE_SNIPPET = _follow_local_storage_snippet(_FAN_FOLLOW_STORAGE_PREFIX)
_FOLLOW_LOCAL_STORAGE_SNIPPET_HATER = _follow_local_storage_snippet(_HATER_FOLLOW_STORAGE_PREFIX)


def ensure_follow_local_storage(html: str, file_path: str) -> str:
    """粉丝/黑粉等有验证的关注页：用 localStorage 记住关注状态（刷新后不丢）。"""
    norm = file_path.replace("\\", "/")
    if "/civilians/" not in norm:
        return html
    if "followStorageKey" in html:
        return html
    if 'id="followModal"' not in html or not re.search(r"var\s+isFollowing\s*=", html):
        return html

    inj_snippet = (
        _FOLLOW_LOCAL_STORAGE_SNIPPET_HATER if "/haters/" in norm else _FOLLOW_LOCAL_STORAGE_SNIPPET
    )

    def _inj_storage_block(s: str) -> str:
        return re.sub(
            r"(var\s+isFollowing\s*=\s*false\s*;)",
            r"\1" + inj_snippet,
            s,
            count=1,
        )

    html2 = _inj_storage_block(html)
    if html2 == html:
        return html

    html = html2
    html = re.sub(
        r"(if\s*\(isFollowing\)\s*\{\s*\n\s+isFollowing\s*=\s*false;\s*\n)(\s+)(render\(\);)",
        r"\1\2persistFollowState();\n\2\3",
        html,
        count=1,
    )
    html = re.sub(
        r"(isFollowing\s*=\s*true;\s*\n)(\s+)(closeModal\(\);\s*\n\s*)(render\(\);)",
        r"\1\2persistFollowState();\n\2\3\4",
        html,
        count=1,
    )

    html, n_chrome = re.subn(
        r"(\n)(\s+)render\(\);(\s*\n\s+setupCommentToggles)",
        r"\1\2loadFollowState();\n\2render();\3",
        html,
        count=1,
    )
    if n_chrome == 0:
        html, n_plain = re.subn(
            r"(\n)(\s+)render\(\);(\s*\n\s+)\}\)\(\);",
            r"\1\2loadFollowState();\n\2render();\3",
            html,
            count=1,
        )
        if n_plain == 0:
            raise ValueError(f"无法在素人页注入 loadFollowState（{file_path}）")

    return html


def ensure_follow_script_iife_closed(html: str, file_path: str) -> str:
    """黑粉页等若 IIFE 未以 })(); 结尾，整块脚本会变成语法错误，关注按钮无任何响应。"""
    norm = file_path.replace("\\", "/")
    if "/civilians/" not in norm:
        return html
    if 'id="followModal"' not in html or "(function () {" not in html:
        return html
    if re.search(r"\}\)\s*\(\)\s*;\s*\r?\n\s*</script>", html):
        return html
    new_html, n = re.subn(
        r"(loadFollowState\(\);\s*\r?\n\s*render\(\);\s*\r?\n)(\s*</script>)",
        r"\1    })();\n\2",
        html,
        count=1,
    )
    return new_html if n else html


def bump_haters_follow_storage_namespace(html: str, file_path: str) -> str:
    """黑粉沿用旧前缀时可能与粉丝共用键；升级到独立前缀并让旧条目失效（等同清空黑粉侧的已关注缓存）。"""
    norm = file_path.replace("\\", "/")
    if "/civilians/haters/" not in norm:
        return html
    old = 'var followStorageKey = "starnet_follow_state::" + profileId;'
    new = f'var followStorageKey = "{_HATER_FOLLOW_STORAGE_PREFIX}" + profileId;'
    if old in html:
        return html.replace(old, new, 1)
    return html


def ensure_follow_modal_portal_to_body(html: str, file_path: str) -> str:
    """`.panel` 用了 overflow:hidden，会裁剪内部的 fixed 遮罩；把 #followModal 挂到 body 上再显示。"""
    norm = file_path.replace("\\", "/")
    if "/civilians/" not in norm or 'id="followModal"' not in html:
        return html
    if "starnet-follow-modal-portal" in html:
        return html
    snippet = (
        "      /* starnet-follow-modal-portal */\n"
        "      if (followModal && followModal.parentElement && followModal.parentElement !== document.body) {\n"
        "        document.body.appendChild(followModal);\n"
        "      }\n"
    )
    new_html, n = re.subn(
        r"(var\s+followModal\s*=\s*document\.getElementById\(\"followModal\"\);\s*\r?\n)",
        r"\1" + snippet,
        html,
        count=1,
    )
    return new_html if n == 1 else html


def ensure_follow_success_feedback(html: str, file_path: str) -> str:
    """关注成功后展示「已成功关注@昵称」与礼花动画。"""
    norm = file_path.replace("\\", "/")
    if ("/social/" not in norm and not norm.startswith("social/")) or 'id="followBtn"' not in html:
        return html

    if "starnet-follow-success-fx" not in html and "</style>" in html:
        css = (
            "\n    /* starnet-follow-success-fx */\n"
            "    .follow-success-toast { position:fixed; left:50%; top:86px; transform:translateX(-50%);"
            " z-index:320; padding:10px 14px; border-radius:999px; border:1px solid #bfe8ca;"
            " background:linear-gradient(90deg,#ecfff3,#f4fffa); color:#1f6d45; font-size:13px;"
            " font-weight:700; box-shadow:0 8px 24px rgba(34,197,94,.18); opacity:0;"
            " animation:followToastIn .2s ease-out forwards; }\n"
            "    .follow-success-toast.leave { animation:followToastOut .28s ease-in forwards; }\n"
            "    .follow-confetti { position:fixed; top:96px; left:50%; width:8px; height:12px; border-radius:2px;"
            " z-index:319; pointer-events:none; animation:followConfetti 900ms ease-out forwards; }\n"
            "    @keyframes followToastIn { from { opacity:0; transform:translate(-50%,-8px);} to { opacity:1; transform:translate(-50%,0);} }\n"
            "    @keyframes followToastOut { from { opacity:1; transform:translate(-50%,0);} to { opacity:0; transform:translate(-50%,-10px);} }\n"
            "    @keyframes followConfetti { from { opacity:1; transform:translate(0,0) rotate(0deg);} to { opacity:0; transform:translate(var(--dx),var(--dy)) rotate(var(--rot));} }\n"
        )
        html = html.replace("</style>", css + "\n  </style>", 1)

    if "function showFollowSuccessFx()" not in html:
        js = (
            "      function showFollowSuccessFx() {\n"
            "        var nameEl = document.querySelector(\".profile .name\");\n"
            "        var nickname = (nameEl && String(nameEl.textContent || \"\").trim()) || \"该用户\";\n"
            "        var toast = document.createElement(\"div\");\n"
            "        toast.className = \"follow-success-toast\";\n"
            "        toast.textContent = \"已成功关注@\" + nickname;\n"
            "        document.body.appendChild(toast);\n"
            "        var colors = [\"#22c55e\", \"#60a5fa\", \"#f59e0b\", \"#a78bfa\", \"#f472b6\"];\n"
            "        for (var i = 0; i < 18; i++) {\n"
            "          var p = document.createElement(\"span\");\n"
            "          p.className = \"follow-confetti\";\n"
            "          p.style.background = colors[i % colors.length];\n"
            "          p.style.setProperty(\"--dx\", (Math.random() * 220 - 110).toFixed(0) + \"px\");\n"
            "          p.style.setProperty(\"--dy\", (80 + Math.random() * 120).toFixed(0) + \"px\");\n"
            "          p.style.setProperty(\"--rot\", (Math.random() * 520 - 260).toFixed(0) + \"deg\");\n"
            "          document.body.appendChild(p);\n"
            "          (function (node) { setTimeout(function () { node.remove(); }, 920); })(p);\n"
            "        }\n"
            "        setTimeout(function () { toast.classList.add(\"leave\"); }, 980);\n"
            "        setTimeout(function () { toast.remove(); }, 1300);\n"
            "      }\n"
        )
        html, n_locked = re.subn(
            r"(var\s+lockedTip\s*=\s*document\.getElementById\(\"lockedTip\"\);\s*\r?\n)",
            r"\1" + js,
            html,
            count=1,
        )
        if n_locked == 0:
            html, _ = re.subn(
                r"(var\s+profileMention\s*=.*?;\s*\r?\n)",
                r"\1" + js,
                html,
                count=1,
            )

    html, _ = re.subn(
        r"(isFollowing\s*=\s*true;\s*\r?\n\s*persistFollowState\(\);\s*\r?\n\s*closeModal\(\);\s*\r?\n\s*render\(\);\s*\r?\n)",
        r"\1        showFollowSuccessFx();\n",
        html,
        count=1,
    )
    html, _ = re.subn(
        r"(isFollowing\s*=\s*true;\s*\r?\n\s*persistFollowState\(\);\s*\r?\n\s*renderFollow\(\);\s*\r?\n)",
        r"\1          showFollowSuccessFx();\n",
        html,
        count=1,
    )
    return html


def update_single_file(profile_id, file_path, profile_row, feed_rows):
    abs_path = ROOT / file_path
    ensure_profile_file(profile_id, file_path, profile_row)

    content = abs_path.read_text(encoding="utf-8")

    handle_slug = slug_from_profile(profile_row, file_path)
    content = apply_profile_identity_urls(content, handle_slug)

    display_name = (profile_row.get("display_name") or "").strip()
    bio = (profile_row.get("bio") or "").strip()
    follow_question = (profile_row.get("follow_question") or "").strip()
    follow_answer = (profile_row.get("follow_answer") or "").strip()

    if display_name:
        content = replace_or_fail(
            r'(<h2 class="name">)(.*?)(</h2>)',
            lambda m: f'{m.group(1)}{html.escape(display_name)}{m.group(3)}',
            content,
            f"未找到 name 区块: {file_path}",
        )

    norm_fp = file_path.replace("\\", "/")

    if "/artists/" in norm_fp:
        content = patch_artist_profile_stats_bar(content, profile_id, len(feed_rows))
    elif "/civilians/fans/" in norm_fp or "/civilians/haters/" in norm_fp:
        content = patch_civilian_profile_stats_bar(content, profile_id, len(feed_rows), file_path)

    if bio:
        content = replace_or_fail(
            r'(<p class="bio">)(.*?)(</p>)',
            lambda m: f'{m.group(1)}{html.escape(bio)}{m.group(3)}',
            content,
            f"未找到 bio 区块: {file_path}",
        )

    if follow_question and 'id="followModal"' in content:
        content = replace_or_fail(
            r'(<div class="modal" id="followModal">.*?<p class="modal-question">)(.*?)(</p>)',
            lambda m: f'{m.group(1)}{html.escape(follow_question)}{m.group(3)}',
            content,
            f"未找到 follow question 区块: {file_path}",
        )

    if follow_answer and re.search(r"var\s+expected\s*=", content):
        js_answer_literal = json.dumps(follow_answer, ensure_ascii=False)
        content = replace_or_fail(
            r'(var expected = )(.*?)(;)',
            lambda m: f"{m.group(1)}{js_answer_literal}{m.group(3)}",
            content,
            f"未找到 expected 变量: {file_path}",
        )

    public_posts_html = []
    protected_posts_html = []

    def _post_rank(row):
        v = str(row.get("post_order") or "").strip()
        try:
            return int(float(v))
        except ValueError:
            return 99999

    def _followers_only_cell(row):
        return str(row.get("followers_only") or "").strip() == "1"

    locked_feed = ('<p class="locked-tip"' in content) and ('id="followModal"' in content)
    merge_followers_into_timeline = locked_feed and any(
        _followers_only_cell(r) for r in feed_rows
    )

    for post_row in sorted(feed_rows, key=_post_rank):
        row_with_file = dict(post_row)
        row_with_file["file"] = file_path
        post_html, is_followers_post = build_post_html(
            profile_id=profile_id, display_name=display_name or "用户", post_row=row_with_file
        )
        if merge_followers_into_timeline and is_followers_post:
            post_html = re.sub(
                r'(<article class="post post-followers")(\s*)>',
                r"\1\2 hidden>",
                post_html,
                count=1,
            )
            public_posts_html.append(post_html)
        elif merge_followers_into_timeline:
            public_posts_html.append(post_html)
        elif is_followers_post:
            protected_posts_html.append(post_html)
        else:
            public_posts_html.append(post_html)

    if feed_rows:
        if merge_followers_into_timeline:
            rebuilt = "".join(public_posts_html)
        elif protected_posts_html:
            rebuilt = (
                "".join(public_posts_html)
                + '<div id="protectedContent" hidden>'
                + "".join(protected_posts_html)
                + "</div>"
            )
        else:
            rebuilt = "".join(public_posts_html)
        if locked_feed:
            content = replace_or_fail(
                r'(<p class="locked-tip" id="lockedTip">.*?</p>)(.*?)(<div class="modal" id="followModal">)',
                lambda m: f"{m.group(1)}\n    {rebuilt}\n  {m.group(3)}",
                content,
                f"未找到可替换的动态区域: {file_path}",
            )
        else:
            content = merge_artist_feed_panel(content, rebuilt)

    _norm_path = file_path.replace("\\", "/")
    if "/civilians/fans/" in _norm_path or "/civilians/haters/" in _norm_path:
        content = ensure_fan_page_chrome(content, file_path)
    if "/civilians/" in file_path.replace("\\", "/"):
        content = ensure_follow_local_storage(content, file_path)
        content = bump_haters_follow_storage_namespace(content, file_path)
        content = ensure_follow_modal_portal_to_body(content, file_path)
        content = ensure_follow_script_iife_closed(content, file_path)
    content = ensure_follow_success_feedback(content, file_path)

    abs_path.write_text(content, encoding="utf-8")


def main():
    if XLSX_PATH.exists():
        profiles = read_sheet_from_xlsx(XLSX_PATH, "profiles")
        feed = []
        for sheet in [
            "feed",
            "posts_艺人",
            "posts_粉丝",
            "posts_黑粉",
            "艺人",
            "素人",
            "黑粉",
        ]:
            feed.extend(read_sheet_from_xlsx(XLSX_PATH, sheet))
    else:
        profiles = read_csv(DATA_DIR / "profiles.csv")
        feed = read_csv(DATA_DIR / "feed.csv")

    if not profiles:
        raise SystemExit("profiles 数据为空，至少需要一条记录。")

    profiles_by_id = {}
    feed_by_profile_id = defaultdict(list)
    for p in profiles:
        profile_id = (p.get("profile_id") or "").strip()
        if profile_id:
            profiles_by_id[profile_id] = p
    for row in feed:
        profile_id = (row.get("profile_id") or "").strip()
        if profile_id:
            feed_by_profile_id[profile_id].append(row)

    updated = 0
    for profile_id, profile in profiles_by_id.items():
        file_path = resolve_file_path(profile_id, profile)
        update_single_file(
            profile_id=profile_id,
            file_path=file_path,
            profile_row=profile,
            feed_rows=feed_by_profile_id.get(profile_id, []),
        )
        updated += 1

    print(f"已更新 {updated} 个主页文件。")


if __name__ == "__main__":
    main()
