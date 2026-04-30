#!/usr/bin/env python3
import csv
import json
import re
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SOCIAL_ROOT = ROOT / "social"
DATA_ROOT = ROOT / "data" / "starnet"
OUT_XLSX = DATA_ROOT / "starnet.xlsx"
PROFILES_CSV = DATA_ROOT / "profiles.csv"


PROFILE_HEADERS = [
    "profile_id",
    "profile_type",
    "file",
    "profile_slug",
    "display_name",
    "bio",
    "follow_question",
    "follow_answer",
]

POST_HEADERS = [
    "profile_id",
    "post_order",
    "time",
    "text",
    "followers_only",
    "image",
    "image_alt",
    "comment_nickname_1",
    "comment_text_1",
    "comment_nickname_2",
    "comment_text_2",
    "comment_nickname_3",
    "comment_text_3",
    "comment_nickname_4",
    "comment_text_4",
    "comment_nickname_5",
    "comment_text_5",
    "comment_nickname_6",
    "comment_text_6",
    "comment_nickname_7",
    "comment_text_7",
    "comment_nickname_8",
    "comment_text_8",
]


def load_profile_id_map():
    result = {}
    if not PROFILES_CSV.exists():
        return result
    with PROFILES_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            file_path = (row.get("file") or "").strip()
            profile_id = (row.get("profile_id") or "").strip()
            if file_path and profile_id:
                result[file_path] = profile_id
    return result


def category_of(file_path):
    if "/artists/" in file_path:
        return "艺人"
    if "/civilians/fans/" in file_path:
        return "粉丝"
    return "黑粉"


def get_follow_answer(content):
    m = re.search(r'var expected = ("(?:[^"\\]|\\.)*")\s*;', content)
    if not m:
        return ""
    try:
        return str(json.loads(m.group(1)))
    except Exception:
        return m.group(1).strip('"')


def normalize_text(node):
    return node.get_text("\n", strip=True).replace("\xa0", " ") if node else ""


def rel_asset_path(profile_file, src):
    if not src:
        return ""
    src = src.strip()
    if re.match(r"^(https?:)?//", src):
        return src
    page_dir = (ROOT / profile_file).resolve().parent
    abs_img = (page_dir / src).resolve()
    try:
        return str(abs_img.relative_to(ROOT)).replace("\\", "/")
    except ValueError:
        return src


def parse_profile(profile_file, profile_id):
    abs_path = ROOT / profile_file
    content = abs_path.read_text(encoding="utf-8")
    soup = BeautifulSoup(content, "html.parser")

    display_name = normalize_text(soup.select_one(".profile .name"))
    bio = normalize_text(soup.select_one(".profile .bio"))

    follow_question = ""
    q_node = soup.select_one("#followModal .modal-question")
    if q_node:
        follow_question = normalize_text(q_node)
        if follow_question.startswith("问题："):
            follow_question = follow_question.replace("问题：", "", 1).strip()

    follow_answer = get_follow_answer(content)

    protected_content = soup.select_one("#protectedContent")
    protected_posts = set(protected_content.select("article.post")) if protected_content else set()

    posts = []
    order = 1
    for article in soup.select("article.post"):
        head = normalize_text(article.select_one(".post-head"))
        time_text = head.split("·", 1)[1].strip() if "·" in head else ""

        p_node = article.find("p")
        text = normalize_text(p_node) if p_node else ""

        img = article.select_one("img.post-image")
        image = rel_asset_path(profile_file, img.get("src", "")) if img else ""
        image_alt = (img.get("alt", "") or "").strip() if img else ""

        followers_only = "1" if article in protected_posts else ""
        post_row = {
            "profile_id": profile_id,
            "post_order": str(order),
            "time": time_text,
            "text": text,
            "followers_only": followers_only,
            "image": image,
            "image_alt": image_alt,
        }
        for i in range(1, 9):
            post_row[f"comment_nickname_{i}"] = ""
            post_row[f"comment_text_{i}"] = ""

        comments = article.select(".comments .comment")
        for idx, c in enumerate(comments[:8], start=1):
            strong = c.find("strong")
            nickname = normalize_text(strong).rstrip("：") if strong else ""
            comment_text = normalize_text(c)
            if strong:
                sn = normalize_text(strong)
                if sn and comment_text.startswith(sn):
                    comment_text = comment_text[len(sn) :].lstrip("：").strip()
            post_row[f"comment_nickname_{idx}"] = nickname
            post_row[f"comment_text_{idx}"] = comment_text

        posts.append(post_row)
        order += 1

    profile_slug = Path(profile_file).stem.replace("starnet-social-", "", 1)
    profile_type = category_of(profile_file)
    profile_row = {
        "profile_id": profile_id,
        "profile_type": profile_type,
        "file": profile_file,
        "profile_slug": profile_slug,
        "display_name": display_name,
        "bio": bio,
        "follow_question": follow_question,
        "follow_answer": follow_answer,
    }
    return profile_row, posts


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def main():
    profile_id_map = load_profile_id_map()
    files = sorted(
        str(p.relative_to(ROOT)).replace("\\", "/")
        for p in SOCIAL_ROOT.rglob("starnet-social-*.html")
    )

    artists_profiles = []
    fans_profiles = []
    haters_profiles = []
    artists_posts = []
    fans_posts = []
    haters_posts = []

    for idx, file_path in enumerate(files, start=1):
        profile_id = profile_id_map.get(file_path, f"P{idx:03d}")
        profile_row, posts = parse_profile(file_path, profile_id)
        cat = category_of(file_path)
        if cat == "艺人":
            artists_profiles.append(profile_row)
            artists_posts.extend(posts)
        elif cat == "粉丝":
            fans_profiles.append(profile_row)
            fans_posts.extend(posts)
        else:
            haters_profiles.append(profile_row)
            haters_posts.extend(posts)

    wb = Workbook()
    ws_profiles = wb.active
    ws_profiles.title = "profiles"
    ws_artist = wb.create_sheet("posts_艺人")
    ws_fan = wb.create_sheet("posts_粉丝")
    ws_hater = wb.create_sheet("posts_黑粉")

    write_sheet(ws_profiles, PROFILE_HEADERS, artists_profiles + fans_profiles + haters_profiles)
    write_sheet(ws_artist, POST_HEADERS, artists_posts)
    write_sheet(ws_fan, POST_HEADERS, fans_posts)
    write_sheet(ws_hater, POST_HEADERS, haters_posts)

    wb.save(OUT_XLSX)
    print(f"Exported to {OUT_XLSX}")
    print(
        f"profiles={len(artists_profiles)+len(fans_profiles)+len(haters_profiles)}, "
        f"artist_posts={len(artists_posts)}, fan_posts={len(fans_posts)}, hater_posts={len(haters_posts)}"
    )


if __name__ == "__main__":
    main()
