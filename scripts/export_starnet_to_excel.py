#!/usr/bin/env python3
import csv
import html
import re
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SOCIAL_DIR = ROOT / "social"
DATA_DIR = ROOT / "data" / "starnet"
OUT_XLSX = DATA_DIR / "starnet.xlsx"
PROFILES_CSV = DATA_DIR / "profiles.csv"


PROFILE_HEADERS = [
    "profile_id",
    "file",
    "display_name",
    "bio",
    "follow_question",
    "follow_answer",
]

FEED_HEADERS = [
    "profile_id",
    "post_order",
    "time",
    "text",
    "followers_only",
    "image",
    "image_alt",
    "comment_nickname_1",
    "comment_text_1",
    "comment_nickname_2",
    "comment_text_2",
    "comment_nickname_3",
    "comment_text_3",
    "comment_nickname_4",
    "comment_text_4",
    "comment_nickname_5",
    "comment_text_5",
    "comment_nickname_6",
    "comment_text_6",
    "comment_nickname_7",
    "comment_text_7",
    "comment_nickname_8",
    "comment_text_8",
]


def read_profile_ids():
    mapping = {}
    if not PROFILES_CSV.exists():
        return mapping
    with PROFILES_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            file_path = (row.get("file") or "").strip()
            profile_id = (row.get("profile_id") or "").strip()
            if file_path and profile_id:
                mapping[file_path] = profile_id
    return mapping


def category_of(path_str):
    if "/artists/" in path_str:
        return "artist"
    if "/civilians/fans/" in path_str:
        return "civilian"
    return "hater"


def strip_html(raw):
    if not raw:
        return ""
    t = raw.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
    t = re.sub(r"</p>\s*<p[^>]*>", "\n", t, flags=re.S)
    t = re.sub(r"<[^>]+>", "", t, flags=re.S)
    return html.unescape(t).strip()


def extract_attr(tag, attr):
    m = re.search(rf'\b{attr}="([^"]*)"', tag)
    return m.group(1).strip() if m else ""


def extract_profile_fields(content):
    def pick(pattern):
        m = re.search(pattern, content, flags=re.S)
        return strip_html(m.group(1)) if m else ""

    display_name = pick(r'<h2 class="name">(.*?)</h2>')
    bio = pick(r'<p class="bio">(.*?)</p>')
    follow_question = pick(
        r'<div class="modal" id="followModal">.*?<p class="modal-question">(.*?)</p>'
    )
    if follow_question.startswith("问题："):
        follow_question = follow_question[len("问题：") :].strip()
    answer_match = re.search(r'var expected = "(.*?)";', content, flags=re.S)
    follow_answer = html.unescape(answer_match.group(1)) if answer_match else ""
    return display_name, bio, follow_question, follow_answer


def extract_comments(article_html):
    comments = []
    block_match = re.search(r'<div class="comments">(.*?)</div>', article_html, flags=re.S)
    if not block_match:
        return comments
    block = block_match.group(1)
    for m in re.finditer(r'<p class="comment">\s*<strong>(.*?)：</strong>(.*?)</p>', block, flags=re.S):
        nickname = strip_html(m.group(1))
        text = strip_html(m.group(2))
        comments.append((nickname, text))
    return comments


def extract_posts(content, file_path):
    protected_ranges = []
    for m in re.finditer(r'<div id="protectedContent"[^>]*>(.*?)</div>', content, flags=re.S):
        protected_ranges.append((m.start(1), m.end(1)))

    def is_protected(start_pos):
        for s, e in protected_ranges:
            if s <= start_pos <= e:
                return "1"
        return ""

    rows = []
    order = 1
    for m in re.finditer(r'(<article class="post[^"]*">.*?</article>)', content, flags=re.S):
        article = m.group(1)
        followers_only = is_protected(m.start(1))
        head = strip_html(re.search(r'<div class="post-head">(.*?)</div>', article, flags=re.S).group(1))
        post_time = ""
        if "·" in head:
            post_time = head.split("·", 1)[1].strip()

        body_match = re.search(r'<div class="post-head">.*?</div>\s*<p>(.*?)</p>', article, flags=re.S)
        post_text = strip_html(body_match.group(1)) if body_match else ""

        image_tag_match = re.search(r'<img\b[^>]*class="[^"]*post-image[^"]*"[^>]*>', article, flags=re.S)
        image = ""
        image_alt = ""
        if image_tag_match:
            tag = image_tag_match.group(0)
            image_src = extract_attr(tag, "src")
            image_alt = extract_attr(tag, "alt")
            if image_src:
                page_dir = (ROOT / file_path).resolve().parent
                image_abs = (page_dir / image_src).resolve()
                try:
                    image = str(image_abs.relative_to(ROOT)).replace("\\", "/")
                except ValueError:
                    image = image_src

        comments = extract_comments(article)
        row = {
            "post_order": str(order),
            "time": post_time,
            "text": post_text,
            "followers_only": followers_only,
            "image": image,
            "image_alt": image_alt,
        }
        for i in range(1, 9):
            row[f"comment_nickname_{i}"] = ""
            row[f"comment_text_{i}"] = ""
        for idx, (nick, text) in enumerate(comments[:8], start=1):
            row[f"comment_nickname_{idx}"] = nick
            row[f"comment_text_{idx}"] = text
        rows.append(row)
        order += 1
    return rows


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def main():
    id_map = read_profile_ids()
    files = sorted(
        (p for p in SOCIAL_DIR.rglob("starnet-social-*.html")),
        key=lambda p: str(p).replace("\\", "/"),
    )

    artists = []
    civilians = []
    haters = []
    feed_artists = []
    feed_civilians = []
    feed_haters = []

    for idx, path in enumerate(files, start=1):
        rel = str(path.relative_to(ROOT)).replace("\\", "/")
        cat = category_of(rel)
        profile_id = id_map.get(rel, f"P{idx:03d}")

        content = path.read_text(encoding="utf-8")
        display_name, bio, follow_question, follow_answer = extract_profile_fields(content)

        profile_row = {
            "profile_id": profile_id,
            "file": rel,
            "display_name": display_name,
            "bio": bio,
            "follow_question": follow_question,
            "follow_answer": follow_answer,
        }
        post_rows = extract_posts(content, rel)
        for r in post_rows:
            r["profile_id"] = profile_id

        if cat == "artist":
            artists.append(profile_row)
            feed_artists.extend(post_rows)
        elif cat == "civilian":
            civilians.append(profile_row)
            feed_civilians.extend(post_rows)
        else:
            haters.append(profile_row)
            feed_haters.extend(post_rows)

    profiles = artists + civilians + haters

    wb = Workbook()
    ws_profiles = wb.active
    ws_profiles.title = "profiles"
    ws_artist = wb.create_sheet("艺人")
    ws_civilian = wb.create_sheet("素人")
    ws_hater = wb.create_sheet("黑粉")

    write_sheet(ws_profiles, PROFILE_HEADERS, profiles)
    write_sheet(ws_artist, FEED_HEADERS, feed_artists)
    write_sheet(ws_civilian, FEED_HEADERS, feed_civilians)
    write_sheet(ws_hater, FEED_HEADERS, feed_haters)

    wb.save(OUT_XLSX)
    print(f"已导出: {OUT_XLSX}")
    print(f"profiles: {len(profiles)}")
    print(f"artist posts: {len(feed_artists)}")
    print(f"civilian posts: {len(feed_civilians)}")
    print(f"hater posts: {len(feed_haters)}")


if __name__ == "__main__":
    main()
