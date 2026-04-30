#!/usr/bin/env python3
"""
Rename StarNet profile HTML to starnet-social-{profile_slug}.html and fix
https://starnet.social/... + @handle across the repo.

Expects data/starnet/starnet.xlsx with a profiles sheet containing at least:
profile_id, profile_type, profile_slug, display_name, file (optional),
and updates the file column to the final path.
"""

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "starnet" / "starnet.xlsx"


def normalize_type(raw):
    t = (raw or "").strip().lower()
    if t in {"artist", "艺人"}:
        return "artist"
    if t in {"fan", "fans", "粉丝", "素人"}:
        return "fan"
    if t in {"hater", "haters", "黑粉"}:
        return "hater"
    return "fan"


def type_to_dir(cat):
    if cat == "artist":
        return "social/artists"
    if cat == "hater":
        return "social/civilians/haters"
    return "social/civilians/fans"


def sanitize_slug(s):
    return re.sub(r"[^a-zA-Z0-9_]+", "_", (s or "").strip()).strip("_")


def derive_slug(row: dict, fallback_id: str) -> str:
    s = sanitize_slug(row.get("profile_slug"))
    if s:
        return s
    f = (row.get("file") or "").strip()
    if f:
        stem = Path(f).stem.removeprefix("starnet-social-")
        stem = sanitize_slug(stem.replace("-", "_")) or sanitize_slug(Path(f).stem)
        if stem:
            return stem
    return sanitize_slug(fallback_id.lower().replace("-", ""))


def stem_from_filename(file_path):
    return Path(file_path).stem.removeprefix("starnet-social-")


def scrape_url_slug(html_text):
    m = re.search(r'<p\s+class="url"[^>]*>([^<]*)</', html_text)
    if not m:
        return ""
    u = (m.group(1) or "").strip()
    if "starnet.social/" not in u:
        return ""
    return u.split("starnet.social/", 1)[-1]


def load_profiles_from_xlsx():
    if not XLSX.exists():
        sys.exit("missing data/starnet/starnet.xlsx")
    wb = load_workbook(XLSX, data_only=False)
    if "profiles" not in wb.sheetnames:
        sys.exit("starnet.xlsx 缺少 profiles 页签")
    ws = wb["profiles"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(x).strip() if x is not None else "" for x in header_row]

    profiles = []
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c + 1).value for c in range(len(headers))]
        if not row_vals or all(v is None or str(v).strip() == "" for v in row_vals):
            continue
        item = {}
        for c, key in enumerate(headers):
            if not key:
                continue
            v = row_vals[c] if c < len(row_vals) else ""
            item[key] = "" if v is None else str(v).strip()
        item["_sheet_row"] = r
        if not item.get("profile_id"):
            continue
        profiles.append(item)

    return wb, ws, headers, profiles


def rebuild_starnet_home_users(text, profiles):
    artists = {}
    for p in profiles:
        if normalize_type(p.get("profile_type")) != "artist":
            continue
        nm = (p.get("display_name") or "").strip()
        short = nm.replace("CONSTELL-", "").strip()
        if short in {"Orion", "Rasal", "Altair", "Regulus"}:
            artists[short] = p

    colors = {
        "Orion": "linear-gradient(135deg,#65b6ff,#6d79ff)",
        "Rasal": "linear-gradient(135deg,#9a8bff,#7ac1ff)",
        "Altair": "linear-gradient(135deg,#72d4ff,#5b8eff)",
        "Regulus": "linear-gradient(135deg,#ffbe75,#ff8c65)",
    }
    avatars = {"Orion": "O", "Rasal": "V", "Altair": "A", "Regulus": "R"}
    order = ["Orion", "Rasal", "Altair", "Regulus"]

    js_lines = []
    for member in order:
        row = artists.get(member)
        if not row:
            continue
        slug = row["_slug"]
        fn = Path(row["file"]).name
        bio_raw = (row.get("bio") or "").strip().replace("\\", "\\\\").replace('"', '\\"')
        if not bio_raw:
            bio_raw = "CONSTELL成员"
        js_lines.append(
            f'        {{ name: "{member}", id: "@{slug}", bio: "{bio_raw}", '
            f'href: "./artists/{fn}", avatar: "{avatars[member]}", color: "{colors[member]}" }}'
        )
    for idx, line in enumerate(js_lines):
        if idx < len(js_lines) - 1:
            js_lines[idx] = line + ","
    block = "\n".join(js_lines)
    m = re.search(r"(var\s+users\s*=\s*\[)([\s\S]*?)(\];\s*\n)", text)
    if not m:
        return text
    return text[: m.start(2)] + "\n" + block + "\n      " + text[m.end(2) :]


def main():
    wb, ws, headers, profiles = load_profiles_from_xlsx()
    file_col = headers.index("file") + 1

    url_pairs = []

    for row in profiles:
        pid = row["profile_id"]
        row["_slug"] = derive_slug(row, pid)
        cat = normalize_type(row.get("profile_type"))
        sheet_file = (ws.cell(row=row["_sheet_row"], column=file_col).value or "").strip()
        row["_old_path"] = sheet_file

        target = f"{type_to_dir(cat)}/starnet-social-{row['_slug']}.html"
        row["file"] = target

        variants = set()
        if sheet_file:
            abs_old = ROOT / sheet_file
            variants.add(stem_from_filename(sheet_file))
            if abs_old.exists():
                variants.add(scrape_url_slug(abs_old.read_text(encoding="utf-8")))
        new_slug = row["_slug"]
        for v in variants:
            v = (v or "").strip()
            if v and v != new_slug:
                url_pairs.append((v, new_slug))

    seen = set()
    dedup = []
    for a, b in url_pairs:
        k = (a, b)
        if k in seen:
            continue
        seen.add(k)
        dedup.append((a, b))
    url_pairs = dedup

    renames = []
    for row in profiles:
        old = row["_old_path"]
        new = row["file"]
        if not old or old == new:
            continue
        renames.append((old.replace("\\", "/"), new.replace("\\", "/")))

    renames.sort(key=lambda x: len(x[0]), reverse=True)
    for old_rel, new_rel in renames:
        old_abs = ROOT / old_rel
        new_abs = ROOT / new_rel
        new_abs.parent.mkdir(parents=True, exist_ok=True)
        if not old_abs.exists():
            continue
        if old_abs.resolve() != new_abs.resolve():
            old_abs.rename(new_abs)

    for row in profiles:
        sr = row["_sheet_row"]
        ws.cell(row=sr, column=file_col, value=row["file"])

    wb.save(XLSX)

    text_files = [
        p for p in ROOT.rglob("*") if p.suffix.lower() in {".html", ".js", ".md", ".csv"}
    ]
    text_files = [p for p in text_files if ".git" not in p.parts]

    for path in text_files:
        if path.resolve() == XLSX.resolve():
            continue
        s = path.read_text(encoding="utf-8")
        orig = s
        for old_rel, new_rel in renames:
            s = s.replace(old_rel, new_rel)
        for old_tail, new_slug in sorted(url_pairs, key=lambda x: len(x[0]), reverse=True):
            s = s.replace(
                f"https://starnet.social/{old_tail}", f"https://starnet.social/{new_slug}"
            )
            s = s.replace(f"@{old_tail}", f"@{new_slug}")
        if s != orig:
            path.write_text(s, encoding="utf-8")

    home = ROOT / "social" / "starnet-home.html"
    if home.exists():
        t = home.read_text(encoding="utf-8")
        t2 = rebuild_starnet_home_users(t, profiles)
        if t2 != t:
            home.write_text(t2, encoding="utf-8")

    print(
        "done: renames=%d url_variants=%d profiles=%d"
        % (len(renames), len(url_pairs), len(profiles))
    )


if __name__ == "__main__":
    main()
